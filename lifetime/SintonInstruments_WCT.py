
import sys
import os
import numpy as np
import matplotlib.pylab as plt
if sys.platform == "win32":
    from win32com.client import Dispatch
import openpyxl as pyxl


def Sinton2014_Settings(File):
    # This grabs the calibration constants for the coil and reference
    # This outputs them as a dictionary

    xlApp = Dispatch("Excel.Application")
    # Set the Excel file name, working directory and path
    # This is encase you don't want to use the directory you are in
    working_dir = r"C:/Users/z3186867/Desktop/"

    File_path = os.path.join(working_dir, File)
    xlBook = xlApp.Workbooks.Open(File_path)

    # make Excel visible (1 is True, 0 is False)
    xlApp.Visible = 0

    # makes a reference to the RawData page
    xlSheet = xlBook.Sheets('Settings')

    Ref = float(xlSheet.Range('C5').Value)
    A = float(xlSheet.Range('C6').Value)
    B = float(xlSheet.Range('C7').Value)
    C = float(xlSheet.Range('C8').Value)
    Air = float(xlSheet.Range('C10').Value)

    xlBook.Close(SaveChanges=0)
    xlApp.Application.Quit()
    Dic = locals()
    for i in ['xlApp', 'xlBook', 'working_dir', 'xlSheet', 'File_path']:
        del Dic[i]
    # print Dic
    return Dic


def Sinton2014_set_UserData(File, settings_dic, Dir=None):
    # This grabs the user entered data about the sample
    # This outputs them as a dictionary
    xlApp = Dispatch("Excel.Application")
    # Set the Excel file name, working directory and path
    # This is encase you don't want to use the directory you are in
    if Dir is None:
        Dir = r"C:\Users\z3186867\Desktop\\"

    File_path = os.path.join(Dir, File)
    # print File_path
    # xlApp.Visible = True
    xlBook = xlApp.Workbooks.Open(File_path)

    # make Excel visible (1 is True, 0 is False)
    xlApp.Visible = 0

    # makes a reference to the RawData page
    xlSheet = xlBook.Sheets('User')

    # Grabbing the data and assigning it a nae
    xlSheet.Range('A6').Value = settings_dic['WaferName']
    xlSheet.Range('B6').Value = settings_dic['Thickness']
    print settings_dic['Resisitivity'], settings_dic
    xlSheet.Range('C6').Value = settings_dic['Resisitivity']
    xlSheet.Range('D6').Value = settings_dic['SampleType']
    xlSheet.Range('H6').Value = settings_dic['AnalysisMode']
    xlSheet.Range('E6').Value = settings_dic['OpticalConstant']

    xlBook.Close(SaveChanges=1)
    xlApp.Application.Quit()
    pass


def _dispatch_sinton2014_extractsserdata(File):
    # This grabs the user entered data about the sample
    # This outputs them as a dictionary
    xlApp = Dispatch("Excel.Application")
    # Set the Excel file name, working directory and path
    # This is encase you don't want to use the directory you are in

    File_path = os.path.join(File)
    # print File_path
    # xlApp.Visible = True
    xlBook = xlApp.Workbooks.Open(File_path)

    # make Excel visible (1 is True, 0 is False)
    xlApp.Visible = 0

    # makes a reference to the RawData page
    xlSheet = xlBook.Sheets('User')

    # Grabbing the data and assigning it a nae
    WaferName = xlSheet.Range('A6').Value.encode('utf8')
    Thickness = float(xlSheet.Range('B6').Value)
    Resisitivity = float(xlSheet.Range('C6').Value)
    Doping = float(xlSheet.Range('J9').Value)
    SampleType = xlSheet.Range('D6').Value.encode('utf8')
    AnalysisMode = xlSheet.Range('H6').Value.encode('utf8')
    OpticalConstant = float(xlSheet.Range('E6').Value)

    # makes a reference to the RawData page
    xlSheet = xlBook.Sheets('Settings')
    A = float(xlSheet.Range('C6').Value)
    B = float(xlSheet.Range('C7').Value)
    C = float(xlSheet.Range('C8').Value)

    RefCell = float(xlSheet.Range('C5').Value)

    xlBook.Close(SaveChanges=0)
    xlApp.Application.Quit()
    # Getting all those names into a dictionary
    Dic = locals()
    # Removing the names i don't want in the dictionary
    # print Dic
    for i in ['xlApp', 'xlBook', 'File_path', 'xlSheet']:
        del Dic[i]
    # print Dic
    return Dic


def _dispatch_Sinton2014_ExtractRawData_noplongerused(File_path):
    # import the Dispatch library, get a reference to an Excel instance
    xlApp = Dispatch("Excel.Application")
    # Set the Excel file name, working directory and path
    # This is encase you don't want to use the directory you are in
    # File_path = os.path.join(working_dir, File)

    # opens the excel book
    xlBook = xlApp.Workbooks.Open(File_path)

    # make Excel visible (1 is True, 0 is False)
    xlApp.Visible = 0

    xlSheet = xlBook.Sheets('Calc')
    Values = np.asarray(xlSheet.Range("A9:I133").Value, dtype=np.float64)

    headers = tuple([[j.encode('utf8') for j in i]
                     for i in xlSheet.Range('A8:I8').Value][0])

    Values2 = np.asarray(xlSheet.Range("O9:P133").Value, dtype=np.float64)
    headers2 = tuple([[j.encode('utf8') for j in i]
                      for i in xlSheet.Range('O8:P8').Value][0])

    # print headers
    # makes a reference to the RawData page
    # xlSheet = xlBook.Sheets('RawData')
    # print Values.dtype.names
    # Get the values from the page
    # Values = np.asarray(xlSheet.Range("A2:G126").Value)
    # headers = tuple([[j.encode('utf8') for j in i]
    #                  for i in xlSheet.Range('A1:G1').Value][0])

    # print type(headers),len(headers)

    xlBook.Close(SaveChanges=0)
    xlApp.Application.Quit()

    Values = np.hstack((Values, Values2))
    headers += headers2

    Out = Values.view(dtype=zip(headers, ['float64'] * len(headers))).copy()

    # Too see it working
    # print Out.dtype.names

    return Out

def extract_raw_data(File_path, Plot=False):
    '''
    This grabs the calculated data from sinton
    Extracts columns A-G of the raw data page.
    Can't extract columns H-I as it is not complete data and this would take some more work
    This outputs the data as a structured array, with the names as the column
    '''

    wb = pyxl.load_workbook(File_path, read_only=True, data_only=True)
    # the command page was only added in the later versions
    if 'Command' in  wb.get_sheet_names():
        data = _openpyxl_Sinton2014_ExtractRawDatadata(wb)
    else:
        try:
            data = _openpyxl_Sinton2014_ExtractRawDatadata(wb)
        except:
            print 'Can''t load this file yet'

    # remove all the nan values from the data
    data = data[~np.isnan(data['Tau (sec)'])]

    if Plot:
        plt.plot(data['Apparent CD'], data['Tau (sec)'], '.-')
        plt.loglog()

    return data

def extract_usr_data(File_path):
    '''
    This grabs the calculated data from sinton
    Extracts columns A-G of the raw data page.
    Can't extract columns H-I as it is not complete data and this would take some more work
    This outputs the data as a structured array, with the names as the column
    '''

    wb = pyxl.load_workbook(File_path, read_only=True, data_only=True)
    # the command page was only added in the later versions
    if 'Command' in  wb.get_sheet_names():
        settings = _openpylx_sinton2014_extractsserdata(wb)
    else:
        try:
            settings = _openpylx_sinton2014_extractsserdata(wb)
        except:
            print 'Can''t load this file yet'

    return settings

def _openpyxl_Sinton2014_ExtractRawDatadata(wb):
    '''
        reads the raw data a sinton WCT-120 spreadsheet form the
        provided instance of the openpylx workbook.
    '''

    # make sure the sheet is in the book
    assert 'Calc' in  wb.get_sheet_names()

    # get the worksheet
    ws = wb.get_sheet_by_name('Calc')

    # get first section of data
    values1 = np.array([[i.value for i in j] for j in ws["A9:I133"]], dtype=np.float64)
    headers1 =  tuple([[j.value.encode('utf8') for j in i] for  i in ws["A8:I8"]][0])

    # get second section of data
    values2 = np.array([[i.value for i in j] for j in ws["O9:P133"]], dtype=np.float64)
    headers2 =  tuple([[j.value.encode('utf8') for j in i] for  i in ws["O8:P8"]][0])

    # form into one array with names
    values = np.hstack((values1, values2))
    headers = headers1 + headers2
    Out = values.view(dtype=zip(headers, ['float64'] * len(headers))).copy()

    # return
    return Out

def _openpylx_sinton2014_extractsserdata(wb):

    # make sure the sheet is in the book
    # get the worksheet
    assert 'User' in  wb.get_sheet_names()
    assert 'Settings' in  wb.get_sheet_names()

    ws = wb.get_sheet_by_name('User')

    # Grabbing the data and assigning it a nae

    user_set =    {
    'wafer_name' : ws['A6'].value.encode('utf8'),
    'thickness' : float(ws['B6'].value),
    'resisitivity' : float(ws['C6'].value),
    'm_resisitivity' : float(ws['C9'].value),
    'doping' : float(ws['J9'].value),
    'sample_type' : ws['D6'].value.encode('utf8'),
    'analysis_mode' : ws['H6'].value.encode('utf8'),
    'optical_constant' : float(ws['E6'].value),
    }

    # makes a reference to the RawData page

    ws = wb.get_sheet_by_name('Settings')

    sys_set = {
    'A' : float(ws['C6'].value),
    'B' : float(ws['C7'].value),
    'C' : float(ws['C8'].value),
    'RefCell' : float(ws['C5'].value),
    }

    # make one dic
    user_set.update(sys_set)

    return user_set


if __name__ == '__main__':
    fnames = [r'/home/shanoot/Temp, while own clod updates/B1-Diff-SiNx.xlsm',
    r'/home/shanoot/Temp, while own clod updates/B2-SiNx-only.xlsm',
    r'/home/shanoot/Temp, while own clod updates/B99-Diff-SiNx.xlsm',
    r'/home/shanoot/Temp, while own clod updates/B100-SiNx-Only.xlsm',
    r'/home/shanoot/Temp, while own clod updates/G1-Diff-SiNx.xlsm',
    r'/home/shanoot/Temp, while own clod updates/G2-SiNx-Only.xlsm',
    r'/home/shanoot/Temp, while own clod updates/G99-Diff-SiNx.xlsm',
    r'/home/shanoot/Temp, while own clod updates/G100-SiNx-only.xlsm',
    ]
    for fname in fnames[:1]:
        print extract_raw_data(fname, True).dtype.names
        # print extract_usr_data(fname),
    plt.show()
