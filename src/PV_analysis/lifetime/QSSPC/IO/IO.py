
import sys
import numpy as np
import openpyxl as pyxl
import scipy.constants as C
import configparser as cp
import re
import numbers

from PV_analysis.lifetime.QSSPC.core import lifetime_QSSPC as LTC

if sys.platform == "win32":
    try:
        from win32com.client import Dispatch
    except:
        pass


from PV_analysis.lifetime.QSSPL.IO import IO as QSSPLIO


def Sinton2014_Settings(File):
    # This grabs the calibration constants for the coil and reference
    # This outputs them as a dictionary
    xlBook = xlApp.Workbooks.Open(File)

    # make Excel visible (1 is True, 0 is False)
    xlApp.Visible = 0

    # makes a reference to the RawData page
    xlSheet = xlBook.Sheets('Settings')

    Ref = float(xlSheet.Range('C5').Value)
    A = float(xlSheet.Range('C6').Value)
    B = float(xlSheet.Range('C7').Value)
    C = float(xlSheet.Range('C8').Value)
    Air = float(xlSheet.Range('C10').Value)

    xlBook.Close(SaveChanges=0)
    xlApp.Application.Quit()
    Dic = locals()
    for i in ['xlApp', 'xlBook', 'working_dir', 'xlSheet', 'File_path']:
        del Dic[i]

    return Dic


def Sinton2014_set_UserData(File_path, settings_dic):
    # This grabs the user entered data about the sample
    # This outputs them as a dictionary
    xlApp = Dispatch("Excel.Application")

    xlBook = xlApp.Workbooks.Open(File_path)

    # make Excel visible (1 is True, 0 is False)
    xlApp.Visible = 0

    # makes a reference to the RawData page
    xlSheet = xlBook.Sheets('User')

    # Grabbing the data and assigning it a nae
    xlSheet.Range('A6').Value = settings_dic['WaferName']
    xlSheet.Range('B6').Value = settings_dic['Thickness']
    xlSheet.Range('C6').Value = settings_dic['Resisitivity']
    xlSheet.Range('D6').Value = settings_dic['SampleType']
    xlSheet.Range('H6').Value = settings_dic['AnalysisMode']
    xlSheet.Range('E6').Value = settings_dic['OpticalConstant']

    xlBook.Close(SaveChanges=1)
    xlApp.Application.Quit()
    pass


def _dispatch_sinton2014_extractsserdata(File_path):
    # This grabs the user entered data about the sample
    # This outputs them as a dictionary
    xlApp = Dispatch("Excel.Application")
    # Set the Excel file name, working directory and path
    # This is encase you don't want to use the directory you are in

    xlBook = xlApp.Workbooks.Open(File_path)

    # make Excel visible (1 is True, 0 is False)
    xlApp.Visible = 0

    # makes a reference to the RawData page
    xlSheet = xlBook.Sheets('User')

    # Grabbing the data and assigning it a nae
    WaferName = xlSheet.Range('A6').Value.encode('utf8')
    Thickness = float(xlSheet.Range('B6').Value)
    Resisitivity = float(xlSheet.Range('C6').Value)
    Doping = float(xlSheet.Range('J9').Value)
    SampleType = xlSheet.Range('D6').Value.encode('utf8')
    AnalysisMode = xlSheet.Range('H6').Value.encode('utf8')
    OpticalConstant = float(xlSheet.Range('E6').Value)

    # makes a reference to the RawData page
    xlSheet = xlBook.Sheets('Settings')
    A = float(xlSheet.Range('C6').Value)
    B = float(xlSheet.Range('C7').Value)
    C = float(xlSheet.Range('C8').Value)

    RefCell = float(xlSheet.Range('C5').Value)

    xlBook.Close(SaveChanges=0)
    xlApp.Application.Quit()
    # Getting all those names into a dictionary
    Dic = locals()
    # Removing the names i don't want in the dictionary

    for i in ['xlApp', 'xlBook', 'File_path', 'xlSheet']:
        del Dic[i]

    return Dic


def _dispatch_Sinton2014_ExtractRawData_noplongerused(File_path):
    # import the Dispatch library, get a reference to an Excel instance
    xlApp = Dispatch("Excel.Application")
    # Set the Excel file name, working directory and path
    # This is encase you don't want to use the directory you are in

    # opens the excel book
    xlBook = xlApp.Workbooks.Open(File_path)

    # make Excel visible (1 is True, 0 is False)
    xlApp.Visible = 0

    xlSheet = xlBook.Sheets('Calc')
    Values = np.asarray(xlSheet.Range("A9:I133").Value, dtype=np.float64)

    headers = tuple([[j.encode('utf8') for j in i]
                     for i in xlSheet.Range('A8:I8').Value][0])

    Values2 = np.asarray(xlSheet.Range("O9:P133").Value, dtype=np.float64)
    headers2 = tuple([[j.encode('utf8') for j in i]
                      for i in xlSheet.Range('O8:P8').Value][0])

    # makes a reference to the RawData page
    # xlSheet = xlBook.Sheets('RawData')

    # Get the values from the page
    # Values = np.asarray(xlSheet.Range("A2:G126").Value)
    # headers = tuple([[j.encode('utf8') for j in i]
    #                  for i in xlSheet.Range('A1:G1').Value][0])

    xlBook.Close(SaveChanges=0)
    xlApp.Application.Quit()

    Values = np.hstack((Values, Values2))
    headers += headers2

    Out = Values.view(dtype=zip(headers, ['float64'] * len(headers))).copy()

    # Too see it working

    return Out


def extract_measurement_data(File_path, Plot=False):
    '''
    This grabs the calculated data from sinton
    Extracts columns A-G of the raw data page.
    Can't extract columns H-I as it is not complete data and this would take
    some more work
    This outputs the data as a structured array, with the names as the column
    '''
    # if its an excle file
    if '.xls' in File_path:
        wb = pyxl.load_workbook(File_path, read_only=True, data_only=True)
        # the command page was only added in the later versions

        if 'Command' in wb.get_sheet_names():
            data = _openpyxl_Sinton2014_ExtractRawDatadata(wb)
        else:
            try:
                data = _openpyxl_Sinton2014_ExtractRawDatadata(wb)
            except:
                print('Can\'t load this file yet')

        # remove all the nan values from the data
        data = data[~np.isnan(data['Tau (sec)'])]

    # if its from the labview version
    elif '.ltr' in File_path:
        data = _labview_sinton2017_extract_raw_data(File_path)

    # TODO
    # is this a place where I should filter the datas generation to make
    # transient work?

    return data


def extract_info(File_path):
    '''
    This grabs the calculated data from sinton
    Extracts columns A-G of the raw data page.
    Can't extract columns H-I as it is not complete data and this would take
    some more work
    This outputs the data as a structured array, with the names as the column
    '''

    # if its an excle file
    if '.xls' in File_path:
        wb = pyxl.load_workbook(File_path, read_only=True, data_only=True)
        # the command page was only added in the later versions
        if 'Command' in wb.get_sheet_names():
            settings = _openpylx_sinton2014_extractsserdata(wb)
        else:
            try:
                settings = _openpylx_sinton2014_extractsserdata(wb)
            except:
                print('Can''t load this file yet')
    # if its from the labview version
    elif '.ltr' in File_path:
        settings = _labview_sinton2017_extract_user_data(File_path)

    return settings


def set_info(File_path, **kwargs):
    '''
    This grabs the calculated data from sinton
    Extracts columns A-G of the raw data page.
    Can't extract columns H-I as it is not complete data and this would take
    some more work
    This outputs the data as a structured array, with the names as the column
    '''

    wb = pyxl.load_workbook(File_path, keep_vba=True)
    # the command page was only added in the later versions
    if 'Command' in wb.get_sheet_names():
        settings = _openpylx_sinton2014_setuserdata(wb, kwargs)
    else:
        try:
            _openpylx_sinton2014_extractsserdata(wb)
        except:
            print('Can\'t access this file type')

    wb.save(File_path.replace('.xlsm', '1.xlsm'))


def _openpyxl_Sinton2014_ExtractRawDatadata(wb):
    '''
        reads the raw and caculated data from the 'Calc' sheet of a  sinton WCT-120 spreadsheet.

        inputs:
            wb:
             instance of a openpylx workbook.
    '''

    # make sure the sheet is in the book
    assert 'Calc' in wb.get_sheet_names()

    # get the worksheet
    ws = wb.get_sheet_by_name('Calc')

    # get first section of data
    values1 = np.array([[i.value for i in j] for j in ws['A9':'I133']],
                       dtype=np.float64)
    headers1 = tuple(
        [[j.value for j in i] for i in ws['A8':'I8']][0])

    def div0check(value):
        if value == '#DIV/0!':
            value = np.inf
        return value
    # get second section of data
    values2 = np.array([[div0check(i.value) for i in j] for j in ws['O9':'Z133']],
                       dtype=np.float64)
    headers2 = tuple(
        [[j.value for j in i] for i in ws['O8':'Z8']][0])

    # form into one array with names
    values = np.hstack((values1, values2))
    headers = headers1 + headers2

    Out = values.view(dtype=list(
        zip(headers, ["float64"] * len(headers)))).copy()

    return Out


def _openpylx_sinton2014_setuserdata(wb, dic):

    # make sure the sheet is in the book
    # get the worksheet
    assert 'User' in wb.get_sheet_names()
    assert 'Settings' in wb.get_sheet_names()

    ws_user = wb.get_sheet_by_name('User')
    ws_settings = wb.get_sheet_by_name('Settings')

    # Grabbing the data and assigning it a nae

    user_dic = {
        'wafer_name': 'A6',
        'thickness': 'B6',
        'resisitivity': 'C6',
        'm_resisitivity': 'C9',
        'doping': 'J9',
        'sample_type': 'D6',
        'analysis_mode': 'H6',
        'optical_constant': 'E6',
        'MCD': 'F6',
        'tau@MCD': 'A9',
        'Voc@1sun': 'K9',
        'J0': 'D9',
        'bulk_tau': 'E9',
    }
    user_set_dic = {
        'A': 'C6',
        'B': 'C7',
        'C': 'C8',
        'RefCell': 'C5',
    }

    for name in dic.keys():
        if name in user_dic.keys():
            ws_user[user_dic[name]] = dic[name]
        else:
            print(name, ' not recognoised')
    # makes a reference to the RawData page

    # save the work book
    pass


def _float_or_none(value):

    if isinstance(value, numbers.Number):
        num = float(value)

    elif type(value) == str:
        try:
            num = float(value.strip('"'))
        except:
            num = None
    else:
        num = None
    return num


def _openpylx_sinton2014_extractsserdata(wb):

    # make sure the sheet is in the book
    # get the worksheet
    assert 'User' in wb.get_sheet_names()
    assert 'Settings' in wb.get_sheet_names()

    ws = wb.get_sheet_by_name('User')

    # Grabbing the data and assigning it a nae
    user_set = {
        'name': ws['A6'].value,
        'thickness': _float_or_none(ws['B6'].value),
        'resisitivity': _float_or_none(ws['C6'].value),
        'm_resisitivity': _float_or_none(ws['C9'].value),
        'doping': _float_or_none(ws['J9'].value),
        'sample_type': ws['D6'].value.encode('utf8'),
        'analysis_mode': ws['H6'].value.encode('utf8'),
        'optical_constant': _float_or_none(ws['E6'].value),
        'absorptance': _float_or_none(ws['E6'].value),
        'MCD': _float_or_none(ws['F6'].value),
        'tau@MCD': _float_or_none(ws['A9'].value),
        'Voc@1sun': _float_or_none(ws['K9'].value),
        'J0': _float_or_none(ws['D9'].value),
        'bulk_tau': _float_or_none(ws['E9'].value),
    }

    # makes a reference to the RawData page

    ws = wb.get_sheet_by_name('Settings')

    sys_set = {
        'A': float(ws['C6'].value),
        'B': float(ws['C7'].value),
        'C': float(ws['C9'].value),
        'RefCell': float(ws['C5'].value),
        'Fit Range': ws['C16'].value,
        'auger_model': ws['C17'].value.encode('utf8'),
    }

    # make one dic
    user_set.update(sys_set)

    ws = wb.get_sheet_by_name('Calc')

    sys_set = {
        'dark_voltage': float(ws['A6'].value)
    }

    user_set.update(sys_set)

    return user_set


def _labview_sinton2017_extract_user_data(fname):
    config = cp.ConfigParser()

    with open(fname, 'rb') as f:
        s = f.read()

    s = re.sub(b'Meas Time.*', b'', s)
    s = re.sub(b'Time of Last Instrument.*', b'', s)
    s = s.decode("utf-8")

    config.read_string(s)

    # Grabbing the data and assigning it a nae
    user_set = {
        'name': config.get(
            'User Inputs', 'Sample Parameters.Sample ID').strip('"'),
        'thickness': _float_or_none(
            config.get('User Inputs', 'Sample Parameters.Thickness (cm)')),
        'resisitivity': _float_or_none(
            config.get('User Inputs', 'Sample Parameters.Base Resistivity (ohm-cm)')),
        'sample_type': config.get(
            'User Inputs', 'Sample Parameters.Sample Type').strip('"').encode("utf-8"),
        'analysis_mode': config.get(
            'User Inputs', 'Analysis Parameters.Analysis Mode').strip('"'),
        'optical_constant': _float_or_none(
            config.get('User Inputs', 'Analysis Parameters.Optical Constant')),
        'absorptance': _float_or_none(
            config.get('User Inputs', 'Analysis Parameters.Optical Constant')),
        'MCD': _float_or_none(
            config.get('User Inputs', 'Analysis Parameters.Lifetime Spec MCD (cm-3)')),
        'J0':  _float_or_none(
            config.get('User Inputs', 'Analysis Parameters.Jo Spec MCD (cm-3)')),
        # 'cell_apeture': config.get(
        #     'User Inputs', 'Analysis Parameters.Ref Cell Aperture Setting').strip('"')
    }

    # extracts all the measurement and system settings
    sys_set = {
        'A': _float_or_none(
            config.get('Calibration Details', 'a')),
        'B': _float_or_none(
            config.get('Calibration Details', 'b')),
        'C': _float_or_none(
            config.get(
                'Zero Instrument Details', 'Avg. Air Voltage')
        ) - _float_or_none(
            config.get('Calibration Details', 'offset \\3d air-c')),
        'air_voltage': _float_or_none(
            config.get('Zero Instrument Details', 'Avg. Air Voltage')),
        'RefCell': _float_or_none(
            config.get('Calibration Details', 'ref cell  (v/sun)')),
        'Fit Range': _float_or_none(
            config.get('Advanced Settings', 'Analysis Parameter Settings.Fit Range')),
        'auger_model': config.get(
            'Advanced Settings', 'Analysis Parameter Settings.Auger Model').strip('"'),
    }

    # make one dic
    user_set.update(sys_set)

    sys_set = {
        'dark_voltage': _float_or_none(config.get('DAQ Output', 'Vdark'))
    }

    user_set.update(sys_set)

    return user_set


def _labview_sinton2017_extract_raw_data(fname):
    config = cp.ConfigParser()
    with open(fname, 'rb') as f:
        s = f.read()
    s = re.sub(b'Meas Time.*', b'', s)
    s = re.sub(b'Time of Last Instrument.*', b'', s)
    s = s.decode("utf-8")

    config.read_string(s)

    time = [float(i) for i in config.get(
        'DAQ Output', 'Time').strip('"').split(' ')[1:]]
    Ref = [float(i) for i in config.get(
        'DAQ Output', 'Ref').strip('"').split(' ')[1:]]
    PC = [float(i) for i in config.get(
        'DAQ Output', 'PC').strip('"').split(' ')[1:]]

    wtype = np.dtype(
        [('Time in s', 'f8'), ('Photovoltage', 'f8'), ('Reference Voltage', 'f8')])
    data = np.empty(len(time), dtype=wtype)
    data['Time in s'] = time
    data['Photovoltage'] = PC
    data['Reference Voltage'] = Ref

    return data


def load_lifetime_sinton(file_path):
    '''
    Loads a Sinton excel and passes it into a lifetime class, with the
    attributes automatically filled. You still need to check that the Sinton
    excel values were correctly choosen.
    '''
    # define the lifetime class
    ltc = LTC()

    # get the measurement data
    data = extract_measurement_data(file_path)
    inf = extract_info(file_path)

    # add raw data
    ltc.time = data['Time in s']
    ltc.PC = data['Photovoltage'] + inf['dark_voltage']
    ltc.dark_voltage = inf['dark_voltage']
    ltc.gen_V = data['Reference Voltage']

    ltc.analysis_options['bgc_side'] = 'end'
    ltc.analysis_options['bgc_value'] = 10
    ltc.analysis_options['bgc_type'] = 'points'

    ltc.coil_constants = {'a': inf.pop('A'),
                          'b': inf.pop('B'),
                          'c': inf.pop('C')}

    ltc.Fs = 0.038 / C.e / inf.pop('RefCell')
    ltc.sample.absorptance = inf.pop('optical_constant')

    # pass externally calculated values
    if 'Minority Carrier Density' in data.dtype.names:
        ltc.sample.nxc = data['Minority Carrier Density']
        ltc.tau = data['Tau (sec)']
        ltc.gen = data['Generation (pairs/s)']

        ltc.intrinsic_tau = np.inf * np.zeros(data['Tau (sec)'].shape[0])

        index = np.isclose(data['Tau (sec)'], 0)
        index *= 1 / data['Tau (sec)'] == data['1/Tau Corrected']

        ltc.intrinsic_tau[~index] = 1. / (
            1. / data['Tau (sec)'][~index] -
            data['1/Tau Corrected'][~index]
        )

        ltc.mobility_sum = data['Conductivity increase'] /\
            data['Apparent CD'] / C.e / inf['thickness']
    # if not cal them
    else:
        pass
        # TODO: needs to calculate the doping to allow calculation of
        # these values from the labview version

    # if there are real numbers for ni use them.
    if 'ni' in data.dtype.names:
        if not np.all(np.isnan(data['ni'])):
            ltc.sample.ni_eff = data['ni']
        else:
            # values taken from the sinton file
            # this is the same value for all versions
            # to 2016.
            ltc.sample.ni = np.sqrt(7.4e19)
            # no BGN valued used
            ltc.sample.ni_eff = 'None'

    # sintons definition of Fs

    if inf['sample_type'] == b'p-type':
        inf['dopant'] = 'boron'
    elif inf['sample_type'] == b'n-type':
        inf['dopant'] = 'phosphorus'

    # sets the dopant type
    inf['dopant_type'] = inf.pop('sample_type').decode("utf-8")

    # Pasa a dic to update atttrs, but turn off warnings
    # for non attributes first
    ltc._warnings = False
    # need to set the dopant-type first, so know what to do when doping is set
    ltc.attrs = {'dopant_type': inf['dopant_type']}
    ltc.attrs = inf
    # turns the warnings back on
    ltc._warnings = True

    # set things specfic to the sinton file
    # make sure the background correction it turned off
    ltc.analysis_options['bgc_type'] = None
    ltc.calibration_method = 'sinton'

    return ltc


def load_raw_voltages(file_path,
                      pc_header='PC_V', time_header='Time_S', gen_header='Generation_V', delimiter='\t'):
    '''
    Loads the raw measurements from a file that is delimited
    It expects the file the be provided with the headers for the columns.
    '''
    # define the lifetime class
    ltc = LTC()

    # get the measurement data
    data = np.genfromtxt(file_path, names=True, delimiter=delimiter)

    # pass to the lifetime class
    ltc.time = data['Time_s']
    ltc.PC = data['PC_V']
    ltc.gen_V = data['Generation_V']

    return ltc


def load_inf_UNSW(fname, ltc_PC, method='python'):
    '''
    inputs:

        fname: (str)
            location of the inf file
        ltc_PC: (class)
            A lifetime class instance for the values to be entered into
        method: (str)
            The method used to load it.

    '''
    settings = QSSPLIO.extract_info(fname, method)

    ltc_PC.coil_constants = {'a': settings['Quad'],
                             'b': settings['Lin'],
                             'c': settings['Const']}

    ltc_PC.sample.temp = settings['Temp']
    ltc_PC.sample.dopant_type = settings['Type'] + '-type'
    ltc_PC.sample.absorptance = 1. - settings['Reflection'] / 100.
    ltc_PC.Fs = settings['Fs']
    ltc_PC.sample.thickness = settings['Thickness']
    ltc_PC.sample.doping = settings['Doping']

    return ltc_PC
